"""
PDF & Image Credit Card Statement Parser
==========================================
Parses credit card PDF statements OR screenshot images (via OCR)
and extracts debit transactions into a formatted XLSX file.

Usage:
    python pdf_parser.py <file_or_folder> [--card CARD_NAME] [--output OUTPUT_FILE]

Examples:
    python pdf_parser.py statement.pdf
    python pdf_parser.py screenshot.png
    python pdf_parser.py statement.pdf --card SBI --output expenses.xlsx
    python pdf_parser.py ./inputs/                 # parse all PDFs & images in folder
"""

import re
import sys
import os
import glob
import argparse
from datetime import datetime

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# OCR backend detection — prefers EasyOCR (pure Python, no system deps),
# falls back to pytesseract if available.
# ---------------------------------------------------------------------------
OCR_BACKEND = None  # Will be set to "easyocr" or "pytesseract" or None

try:
    import pytesseract
    from PIL import Image, ImageFilter, ImageOps
    OCR_BACKEND = "pytesseract"
except ImportError:
    pass

try:
    import easyocr
    OCR_BACKEND = "easyocr"
except ImportError:
    pass

# PIL is also useful for easyocr preprocessing
try:
    from PIL import Image, ImageFilter, ImageOps
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

HAS_OCR = OCR_BACKEND is not None


# ---------------------------------------------------------------------------
# Supported file extensions
# ---------------------------------------------------------------------------
PDF_EXTENSIONS = {".pdf"}
IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".webp"}
ALL_EXTENSIONS = PDF_EXTENSIONS | IMAGE_EXTENSIONS


# ---------------------------------------------------------------------------
# Tesseract path auto-detection (Windows) — only used if pytesseract backend
# ---------------------------------------------------------------------------
def _detect_tesseract():
    """Try to find tesseract binary on Windows if not already on PATH."""
    if OCR_BACKEND != "pytesseract":
        return
    candidates = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        os.path.expanduser(r"~\AppData\Local\Programs\Tesseract-OCR\tesseract.exe"),
        os.path.expanduser(r"~\AppData\Local\Tesseract-OCR\tesseract.exe"),
    ]
    for path in candidates:
        if os.path.isfile(path):
            pytesseract.pytesseract.tesseract_cmd = path
            return

_detect_tesseract()


# ---------------------------------------------------------------------------
# Regex patterns for parsing credit card statement lines
# ---------------------------------------------------------------------------

# Matches a transaction line:
#   Date (DD Mon YY)  Description  Amount (with commas/decimals)  C/D
TRANSACTION_RE = re.compile(
    r"^(\d{2}\s+\w{3}\s+\d{2})\s+"   # date: 03 Jun 26
    r"(.+?)\s+"                        # description (non-greedy)
    r"([\d,]+\.\d{2})\s+"             # amount: 1,912.00
    r"([CD])\s*$"                      # type: C or D
)

# OCR can sometimes misread characters, or completely miss the isolated C/D at the edge, so we use a more lenient pattern
TRANSACTION_OCR_RE = re.compile(
    r"^(\d{2}\s+\w{3}\s+\d{2})\s+"    # date: 03 Jun 26
    r"(.+?)\s+"                         # description (non-greedy)
    r"([\d,]+\.\d{2})\s*"              # amount: 1,912.00
    r"([CDcd])?\s*$"                    # type: C or D (optional for OCR)
)

# Format 2 (Alternate Statement Type):
# 01/06/2026| 00:00  ETERNAL LIMITEDGURGAON  - 8  + ₹ 468.84
FORMAT2_RE = re.compile(
    r"^(\d{2}/\d{2}/\d{4}\|\s*\d{2}:\d{2})\s+"      # date and time (cleaned by _fix_ocr_errors)
    r"(.+?)\s+"                                     # description (non-greedy)
    r"(?:(?:[+-]\s*)?\d+\s+)?"                      # optional rewards (ignored)
    r"([+-]?)\s*(?:₹)?\s*([\d,]+\.\d{2})\s*$"       # sign and amount
)

# Lines to skip (headers, footers, section titles, etc.)
SKIP_PATTERNS = [
    re.compile(r"(?i)^\s*date\s+transaction\s+details", re.IGNORECASE),
    re.compile(r"(?i)^\s*for\s+statement\s+period", re.IGNORECASE),
    re.compile(r"(?i)^\s*amount\s*[\(\[]", re.IGNORECASE),
    re.compile(r"(?i)^\s*transactions?\s+for\s+", re.IGNORECASE),
    re.compile(r"(?i)^\s*page\s+\d+", re.IGNORECASE),
    re.compile(r"(?i)^\s*statement\s+summary", re.IGNORECASE),
    re.compile(r"(?i)^\s*credit\s+limit", re.IGNORECASE),
    re.compile(r"(?i)^\s*total\s+", re.IGNORECASE),
    re.compile(r"(?i)^\s*opening\s+balance", re.IGNORECASE),
    re.compile(r"(?i)^\s*closing\s+balance", re.IGNORECASE),
    re.compile(r"(?i)^\s*amount\s*\(", re.IGNORECASE),
    re.compile(r"^\s*$"),  # blank lines
]


def should_skip_line(line: str) -> bool:
    """Check if a line should be skipped (headers, totals, etc.)."""
    for pattern in SKIP_PATTERNS:
        if pattern.search(line):
            return True
    return False


def parse_date(date_str: str) -> datetime:
    """Parse date string like '03 Jun 26' into a datetime object."""
    # Handle 2-digit year: assume 2000s
    return datetime.strptime(date_str, "%d %b %y")


def parse_date_format2(date_str: str) -> datetime:
    """Parse date string like '01/06/2026| 00:00' into a datetime object."""
    date_part = date_str.split('|')[0].strip()
    return datetime.strptime(date_part, "%d/%m/%Y")


def clean_amount(amount_str: str) -> float:
    """Convert amount string like '1,912.00' to float."""
    return float(amount_str.replace(",", ""))


def simplify_description(desc: str) -> str:
    """
    Simplify the transaction description into a cleaner remark.
    Removes city/state codes, extra whitespace, and common noise.
    """
    desc = desc.strip()
    
    # Remove trailing state/country codes (2-3 letter codes at the end)
    desc = re.sub(r"\s+[A-Z]{2,3}\s*$", "", desc)
    # Remove trailing city names followed by state codes
    desc = re.sub(r"\s+(BANGALORE|BENGALURU|DELHI|NEW\s+DELHI|MUMBAI|NOIDA|JAIPUR|CHENNAI|HYDERABAD|KOLKATA|PUNE|GURUGRAM|GURGAON)\s*.*$", "", desc, flags=re.IGNORECASE)
    # Remove "Bangalore" or "bangalore" (case-insensitive) even without trailing code
    desc = re.sub(r"\s+Bangalore\s*$", "", desc, flags=re.IGNORECASE)
    
    # Clean up specific merchant patterns
    desc = re.sub(r"\*", " ", desc)              # Replace * with space
    desc = re.sub(r"\s{2,}", " ", desc)           # Collapse multiple spaces
    desc = desc.strip()
    
    # Common merchant name simplifications
    merchant_map = {
        r"(?i)amazon\s*seller\s*services": "Amazon",
        r"(?i)zepto\s*marketplace\s*pri": "Zepto",
        r"(?i)beminimalist": "Minimalist",
        r"(?i)uniqlo\s*india\s*private": "Uniqlo",
        r"(?i)ptm\s*reliance\s*retail\s*l": "Reliance Retail",
        r"(?i)raz\s*blue\s*tokai\s*coffee": "Blue Tokai Coffee",
        r"(?i)asspl": "Amazon",
        r"(?i)myntra\s*designs\s*private": "Myntra",
        r"(?i)pay\s*cma\s*equipments": "CMA Equipments",
        r"(?i)rsp\s*carbontree\s*cloth": "Carbontree",
        r"(?i)rsp\s*blink\s*commerce": "Blinkit",
        r"(?i)bling\s*queen": "Bling Queen",
        r"(?i)orbgen\s*technologies": "Orbgen Technologies",
        r"(?i)eternal\s*limited": "Eternal Limited",
        r"(?i)swiggy": "Swiggy",
        r"(?i)zomato": "Zomato",
        r"(?i)flipkart": "Flipkart",
        r"(?i)bigbasket": "BigBasket",
        r"(?i)uber\s*india": "Uber",
        r"(?i)ola\s*money": "Ola",
        r"(?i)paytm": "Paytm",
        r"(?i)netflix": "Netflix",
        r"(?i)spotify": "Spotify",
        r"(?i)google\s*\*": "Google",
        r"(?i)apple\.com": "Apple",
        r"(?i)instamart": "Swiggy Instamart",
        r"(?i)airport\s*lounge": "Airport Lounge",
        r"(?i)bppy\s*cc\s*payment": "CC Payment",
    }
    
    for pattern, replacement in merchant_map.items():
        if re.search(pattern, desc):
            return replacement
    
    return desc


# ---------------------------------------------------------------------------
# Shared line parsing
# ---------------------------------------------------------------------------

def parse_lines_to_transactions(lines: list[str], use_ocr_regex: bool = False) -> list[dict]:
    """
    Parse a list of text lines into transaction dicts.
    Shared by both PDF and image extraction paths.
    """
    pattern = TRANSACTION_OCR_RE if use_ocr_regex else TRANSACTION_RE
    transactions = []
    
    for line in lines:
        line = line.strip()
        
        if should_skip_line(line):
            continue
        
        match = pattern.match(line)
        
        # Check against Format 2 if standard format didn't match
        match2 = None
        if not match:
            match2 = FORMAT2_RE.match(line)
            
        if match:
            date_str = match.group(1)
            description = match.group(2).strip()
            amount_str = match.group(3)
            if match.group(4):
                txn_type = match.group(4).upper()  # Normalize to uppercase
            else:
                # Fallback inference if OCR missed the C or D completely
                desc_lower = description.lower()
                if any(word in desc_lower for word in ["payment", "cashback", "credit", "refund", "reversal"]):
                    txn_type = "C"
                else:
                    txn_type = "D"
            
            try:
                amt = clean_amount(amount_str)
                if txn_type == "C":
                    amt = -amt
                    
                txn = {
                    "date": parse_date(date_str),
                    "description": description,
                    "amount": amt,
                    "type": txn_type,   # C = Credit, D = Debit
                    "remark": simplify_description(description),
                }
                transactions.append(txn)
            except ValueError:
                continue
                
        elif match2:
            date_str = match2.group(1)
            description = match2.group(2).strip()
            sign = match2.group(3).strip()
            amount_str = match2.group(4)
            
            # '+' indicates a Credit, otherwise Debit
            txn_type = "C" if sign == "+" else "D"
            
            try:
                amt = clean_amount(amount_str)
                if txn_type == "C":
                    amt = -amt
                    
                txn = {
                    "date": parse_date_format2(date_str),
                    "description": description,
                    "amount": amt,
                    "type": txn_type,
                    "remark": simplify_description(description),
                }
                transactions.append(txn)
            except ValueError:
                continue
    
    return transactions


# ---------------------------------------------------------------------------
# PDF extraction
# ---------------------------------------------------------------------------

def extract_transactions_from_pdf(pdf_path: str, password: str = None) -> list[dict]:
    """
    Extract all transactions from a credit card PDF statement.
    Supports password-protected PDFs.
    Returns a list of dicts with keys: date, description, amount, type, remark
    """
    all_lines = []
    
    open_kwargs = {}
    if password:
        open_kwargs["password"] = password
    
    with pdfplumber.open(pdf_path, **open_kwargs) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            all_lines.extend(text.split("\n"))
    
    return parse_lines_to_transactions(all_lines, use_ocr_regex=False)


# ---------------------------------------------------------------------------
# Image / OCR extraction
# ---------------------------------------------------------------------------

# Lazy-loaded EasyOCR reader (heavy init, reuse across calls)
_easyocr_reader = None


def _get_easyocr_reader():
    """Get or create a cached EasyOCR reader instance."""
    global _easyocr_reader
    if _easyocr_reader is None:
        print("       Initializing EasyOCR (first run downloads models)...")
        _easyocr_reader = easyocr.Reader(["en"], gpu=False)
    return _easyocr_reader


def _ocr_with_easyocr(image_path: str) -> str:
    """Run OCR on an image using EasyOCR and return reconstructed text lines."""
    reader = _get_easyocr_reader()
    
    # EasyOCR returns list of (bbox, text, confidence)
    # Use min_size=1 to avoid dropping small standalone characters like 'C' or 'D' on the far edge
    results = reader.readtext(image_path, detail=1, paragraph=False, min_size=1, text_threshold=0.1)
    
    if not results:
        return ""
    
    # Sort results by vertical position (top of bounding box), then horizontal
    # Each result: (bbox, text, confidence)
    # bbox is [[x1,y1],[x2,y2],[x3,y3],[x4,y4]]
    sorted_results = sorted(results, key=lambda r: (r[0][0][1], r[0][0][0]))
    
    # Group into lines by Y-coordinate proximity
    lines = []
    current_line = []
    current_y = None
    y_threshold = 15  # pixels — texts within this vertical distance are same line
    
    for bbox, text, conf in sorted_results:
        top_y = (bbox[0][1] + bbox[3][1]) / 2  # average Y of left edge
        
        if current_y is None or abs(top_y - current_y) <= y_threshold:
            current_line.append((bbox[0][0], text))  # (x_position, text)
            current_y = top_y if current_y is None else (current_y + top_y) / 2
        else:
            # New line — flush current
            current_line.sort(key=lambda x: x[0])  # sort by x position
            lines.append(" ".join(t for _, t in current_line))
            current_line = [(bbox[0][0], text)]
            current_y = top_y
    
    # Flush last line
    if current_line:
        current_line.sort(key=lambda x: x[0])
        lines.append(" ".join(t for _, t in current_line))
    
    return "\n".join(lines)


def _ocr_with_pytesseract(image_path: str) -> str:
    """Run OCR on an image using pytesseract."""
    img = Image.open(image_path)
    
    # Preprocess: grayscale, contrast, sharpen, upscale, binarise
    img = img.convert("L")
    img = ImageOps.autocontrast(img, cutoff=1)
    img = img.filter(ImageFilter.SHARPEN)
    
    width, height = img.size
    if width < 1500:
        scale = 1500 / width
        img = img.resize((int(width * scale), int(height * scale)), Image.LANCZOS)
    
    img = img.point(lambda p: 255 if p > 140 else 0, "1")
    
    custom_config = r"--oem 3 --psm 6"
    return pytesseract.image_to_string(img, config=custom_config)


def _fix_ocr_errors(line: str) -> str:
    """Fix common OCR misread characters in credit card statement lines."""
    month_fixes = {
        "3un": "Jun", "3an": "Jan", "Dun": "Jun",
        "Jum": "Jun", "Jui": "Jul", "Aup": "Aug",
        "0ct": "Oct", "Qct": "Oct",
    }
    for wrong, right in month_fixes.items():
        line = line.replace(wrong, right)
    
    # Fix 'O' misread as '0' in amount at end of line (Format 1)
    line = re.sub(r"(\d)O(\s+[CDcd]\s*$)", r"\g<1>0\2", line)
    line = re.sub(r"O(\d{2}\s+[CDcd]\s*$)", r"0\1", line)
    
    # Format 2 specific OCR fixes
    # 1. Fix date/time separator (e.g. '01/06/20261 00.00' -> '01/06/2026| 00:00')
    line = re.sub(r"^(\d{2}/\d{2}/\d{4})[|1lI\s/,]+(\d{2})[.:\s]?(\d{2})\d*\b", r"\1| \2:\3", line)
    
    # 2. Fix Rupee symbol misreads ('{', '<', '?', 'F', 'R', 'z' followed by amount)
    line = re.sub(r"[{<?FRz]\s*([\d,.]+\.\d{2})\b", r"₹ \1", line)
    
    # 3. Fix '+' misread as '3' or omitted entirely before amounts (e.g. '3 24,115.00' -> '+ ₹ 24,115.00')
    line = re.sub(r"\b3\s+(?:₹\s*)?([\d,.]+\.\d{2})\b", r"+ ₹ \1", line)
    line = re.sub(r"\+\s*(?:₹\s*)?([\d,.]+\.\d{2})\b", r"+ ₹ \1", line)
    
    # 4. Fix commas misread as dots in amounts (e.g. '11.950.00' -> '11,950.00')
    line = re.sub(r"(\d)\.(\d{3})\.(\d{2})\b", r"\1,\2.\3", line)
    line = re.sub(r"(\d)\.(\d{3})\,(\d{2})\b", r"\1,\2.\3", line)
    
    return line


def extract_transactions_from_image(image_path: str, ocr_engine: str = "easyocr") -> list[dict]:
    """
    Extract transactions from a screenshot image using OCR.
    Uses EasyOCR (preferred) or pytesseract as fallback.
    """
    if not HAS_OCR:
        print("  -> Error: No OCR backend available.")
        print("     Install EasyOCR:    pip install easyocr")
        print("     Or pytesseract:     pip install pytesseract Pillow")
        print("       + Tesseract OCR:  https://github.com/UB-Mannheim/tesseract/wiki")
        return []
    
    # Run OCR with the requested backend (if available)
    if ocr_engine == "easyocr" and OCR_BACKEND == "easyocr":
        text = _ocr_with_easyocr(image_path)
    else:
        text = _ocr_with_pytesseract(image_path)
    
    if not text or not text.strip():
        print("  -> Warning: OCR returned empty text. Try a higher resolution image.")
        return []
    
    # Fix common OCR errors and parse
    lines = text.split("\n")
    fixed_lines = [_fix_ocr_errors(line) for line in lines]
    
    # Debug: print OCR output
    print("  -> OCR extracted lines:")
    for i, line in enumerate(fixed_lines):
        f1_match = "F1" if TRANSACTION_OCR_RE.match(line) else "  "
        f2_match = "F2" if FORMAT2_RE.match(line) else "  "
        print(f"     [{f1_match}|{f2_match}] {repr(line)}")
    
    return parse_lines_to_transactions(fixed_lines, use_ocr_regex=True)


# ---------------------------------------------------------------------------
# Unified file extraction
# ---------------------------------------------------------------------------

def extract_transactions_from_file(file_path: str, ocr_engine: str = "easyocr") -> list[dict]:
    """
    Extract transactions from a file (PDF or image).
    Auto-detects file type by extension.
    """
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext in PDF_EXTENSIONS:
        return extract_transactions_from_pdf(file_path)
    elif ext in IMAGE_EXTENSIONS:
        return extract_transactions_from_image(file_path, ocr_engine=ocr_engine)
    else:
        print(f"  -> Skipping unsupported file type: {ext}")
        return []


# ---------------------------------------------------------------------------
# XLSX output
# ---------------------------------------------------------------------------

def write_to_xlsx(transactions: list[dict], output_path: str, card_name: str = "SBI", style: int = 1):
    """
    Write extracted transactions to a formatted XLSX file.
    
    Columns:
        A - Date
        B - Amount
        C - My Share (blank - user input)
        D - Nitt Share (formula: =IF(B<row><>0, B<row>-C<row>, ""))
        E - Remarks
        F - Category (blank - user input)
        G - Card
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Credit_{card_name}"
    
    # ---- Styles ----
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    if style == 2:
        header_fill = PatternFill(start_color="5C739C", end_color="5C739C", fill_type="solid")
    else:
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    data_font = Font(name="Calibri", size=11)
    amount_format = "#,##0.00"
    date_format = "DD/MM/YYYY"
    
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )
    
    # Column C highlight (blank for user input) - only applied in style 1
    col_c_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Alternating row colors for style 2
    row_fill_light = PatternFill(start_color="E9EEF4", end_color="E9EEF4", fill_type="solid")
    row_fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    # ---- Headers (Row 2 to match the screenshot pattern) ----
    headers = ["Date", "Amount", "My Share", "Nitt Share", "Remarks", "Category", "Card"]
    header_row = 2
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # ---- Column widths ----
    col_widths = {
        "A": 22,   # Date
        "B": 14,   # Amount
        "C": 14,   # My Share
        "D": 14,   # Nitt Share
        "E": 30,   # Remarks
        "F": 16,   # Category
        "G": 10,   # Card
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # ---- Data rows (starting from row 3) ----
    
    for row_idx, txn in enumerate(transactions, start=3):
        row_fill = row_fill_white
        if style == 2:
            row_fill = row_fill_light if row_idx % 2 == 1 else row_fill_white
            
        # Column A: Date
        date_cell = ws.cell(row=row_idx, column=1, value=txn["date"])
        date_cell.number_format = date_format
        date_cell.font = data_font
        date_cell.border = thin_border
        
        # Column B: Amount
        amount_cell = ws.cell(row=row_idx, column=2, value=txn["amount"])
        amount_cell.number_format = amount_format
        amount_cell.font = data_font
        amount_cell.border = thin_border
        
        # Column C: My Share (blank - user input)
        share_cell = ws.cell(row=row_idx, column=3)
        share_cell.fill = col_c_fill if style == 1 else row_fill
        share_cell.number_format = amount_format
        share_cell.font = data_font
        share_cell.border = thin_border
        
        # Column D: Nitt Share (formula)
        formula = f'=IF(B{row_idx}<>0,B{row_idx}-C{row_idx},"")'
        nitt_cell = ws.cell(row=row_idx, column=4, value=formula)
        nitt_cell.number_format = amount_format
        nitt_cell.font = data_font
        nitt_cell.border = thin_border
        
        # Column E: Remarks
        remark_cell = ws.cell(row=row_idx, column=5, value=txn["remark"])
        remark_cell.font = data_font
        remark_cell.border = thin_border
        
        # Column F: Category (blank - user input)
        cat_cell = ws.cell(row=row_idx, column=6)
        cat_cell.font = data_font
        cat_cell.border = thin_border
        
        # Column G: Card
        card_cell = ws.cell(row=row_idx, column=7, value=card_name)
        card_cell.font = data_font
        card_cell.border = thin_border
        
        # Apply row fill for style 2 across all cells in the row
        if style == 2:
            date_cell.fill = row_fill
            amount_cell.fill = row_fill
            nitt_cell.fill = row_fill
            remark_cell.fill = row_fill
            cat_cell.fill = row_fill
            card_cell.fill = row_fill
    
    # ---- Auto-filter on header row ----
    last_row = header_row + len(transactions)
    ws.auto_filter.ref = f"A{header_row}:G{last_row}"
    
    # ---- Freeze panes (freeze below header) ----
    ws.freeze_panes = f"A{header_row + 1}"
    
    wb.save(output_path)
    return len(transactions)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def collect_input_files(input_path: str) -> list[str]:
    """
    Collect all supported files from a path (single file or directory).
    Returns sorted list of file paths.
    """
    if os.path.isdir(input_path):
        files = []
        for ext in ALL_EXTENSIONS:
            files.extend(glob.glob(os.path.join(input_path, f"*{ext}")))
            files.extend(glob.glob(os.path.join(input_path, f"*{ext.upper()}")))
        files = sorted(set(files))
        return files
    elif os.path.isfile(input_path):
        ext = os.path.splitext(input_path)[1].lower()
        if ext in ALL_EXTENSIONS:
            return [input_path]
        else:
            print(f"Error: Unsupported file type '{ext}'")
            print(f"Supported: {', '.join(sorted(ALL_EXTENSIONS))}")
            sys.exit(1)
    else:
        print(f"Error: '{input_path}' is not a valid file or directory.")
        sys.exit(1)


def main():
    parser = argparse.ArgumentParser(
        description="Parse credit card PDF statements or screenshot images and export to XLSX.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python pdf_parser.py statement.pdf
  python pdf_parser.py screenshot.png
  python pdf_parser.py screenshot.jpg --card HDFC --output expenses.xlsx
  python pdf_parser.py ./inputs/            # parse all PDFs & images in folder

Supported file types:
  PDF:    .pdf
  Images: .png, .jpg, .jpeg, .bmp, .tiff, .tif, .webp
        """
    )
    parser.add_argument(
        "input",
        help="Path to a PDF/image file or a folder containing them."
    )
    parser.add_argument(
        "--card", "-c",
        default="SBI",
        help="Card name to write in the 'Card' column (default: SBI)."
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        help="Output XLSX file path (default: <input_name>_parsed.xlsx)."
    )
    
    args = parser.parse_args()
    input_path = args.input
    card_name = args.card
    
    # Show OCR backend info
    if OCR_BACKEND:
        print(f"OCR backend: {OCR_BACKEND}")
    else:
        print("OCR backend: none (image parsing unavailable)")
    
    # Collect input files
    input_files = collect_input_files(input_path)
    if not input_files:
        print(f"Error: No supported files found in '{input_path}'")
        print(f"Supported: {', '.join(sorted(ALL_EXTENSIONS))}")
        sys.exit(1)
    
    pdf_count = sum(1 for f in input_files if os.path.splitext(f)[1].lower() in PDF_EXTENSIONS)
    img_count = sum(1 for f in input_files if os.path.splitext(f)[1].lower() in IMAGE_EXTENSIONS)
    print(f"Found {len(input_files)} file(s): {pdf_count} PDF(s), {img_count} image(s)")
    
    if img_count > 0 and not HAS_OCR:
        print("\nWarning: Image files found but no OCR backend available.")
        print("  Install EasyOCR (recommended): pip install easyocr")
        print("  Or pytesseract:                pip install pytesseract Pillow")
        print("  Image files will be skipped.\n")
    
    # Parse all files
    all_transactions = []
    for file_path in input_files:
        ext = os.path.splitext(file_path)[1].lower()
        source_type = "PDF" if ext in PDF_EXTENSIONS else "IMG"
        print(f"[{source_type}] Parsing: {os.path.basename(file_path)}")
        
        try:
            txns = extract_transactions_from_file(file_path)
            debit_count = sum(1 for t in txns if t["type"] == "D")
            print(f"       -> {len(txns)} transactions found ({debit_count} debits)")
            all_transactions.extend(txns)
        except Exception as e:
            print(f"       -> Error: {e}")
    
    if not all_transactions:
        print("\nNo transactions found in any file.")
        sys.exit(1)
    
    # Sort by date
    all_transactions.sort(key=lambda t: t["date"])
    
    # Determine output path
    if args.output:
        output_path = args.output
    else:
        if len(input_files) == 1:
            base_name = os.path.splitext(os.path.basename(input_files[0]))[0]
        else:
            base_name = "cc_statement"
        output_dir = os.path.dirname(input_path) or "."
        output_path = os.path.join(output_dir, f"{base_name}_parsed.xlsx")
    
    # Write XLSX
    count = write_to_xlsx(all_transactions, output_path, card_name)
    
    total = len(all_transactions)
    skipped = total - count
    print(f"\n{'='*50}")
    print(f"Done! Wrote {count} debit transactions to '{output_path}'")
    if skipped > 0:
        print(f"  - Skipped {skipped} credit transaction(s)")
    print(f"  - Column C ('My Share') is blank for your input")
    print(f"  - Column D ('Nitt Share') has formula =IF(B<row><>0, B<row>-C<row>, \"\")")
    print(f"{'='*50}")


if __name__ == "__main__":
    main()
